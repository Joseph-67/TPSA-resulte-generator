import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from datetime import datetime

# Your data
from classes.grade.subject import subjects
from classes.grade.gradeOneDiamond import students_data, class_data

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# Styles
header_font = Font(name="Times New Roman", bold=True, size=32)
sub_header_font = Font(name="Times New Roman", bold=True, size=16)
title_font = Font(name="Times New Roman", bold=True, size=14)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
gray_fill = PatternFill("solid", fgColor="DDDDDD")
section_fill = PatternFill("solid", fgColor="DDDDDD")

# Column widths
col_widths = {"A": 40.71, "B": 15, "C": 15, "D": 25, "E" : 15, "F": 15, "G": 15, "H": 15, "I": 15, "J": 10, "K" : 10, "L": 20, "M": 5}

def safe_sheet_name(base, term):
    name = f"{base[:18]}_{term}"
    return name[:31]  # Excel limit

def create_term_sheet(ws, student, term_name, term_num, show_prev=True):
    # Page setup
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    ws.page_setup.horizontalCentered = True

    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    row = 1

    # Header
    ws.merge_cells('A1:M1')
    ws['A1'] = "TOPSTEPS ACADEMY"
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 38

    ws.merge_cells('A2:M2')
    ws['A2'] = f"{term_name.upper()} TERM 2025/2026 SESSION ACADEMIC REPORT"
    ws['A2'].font = sub_header_font
    ws['A2'].alignment = center_align
    ws['A2'].fill = gray_fill
    ws.row_dimensions[2].height = 25

    if os.path.exists("tpsa_logo.png"):
        logo = Image("tpsa_logo.png")
        logo.width = logo.height = 80
        ws.add_image(logo, "A1")

    row = 3

    # Student Info

    ws.merge_cells('B3:F3')
    ws.merge_cells('G3:H3')
    ws.merge_cells('I3:M3')
    ws.merge_cells('B4:C4')
    ws.merge_cells('E4:F4')
    ws.merge_cells('G4:H4')
    ws.merge_cells('I4:M4')
    ws.merge_cells('B5:M5')
    ws.merge_cells('A7:M7')
    

    ws[f'A{row}'] = "Name:"; ws[f'B{row}'] = student["Name"].upper()
    ws[f'G{row}'] = "Adm No:"; ws[f'I{row}'] = student["Admission No"]
    ws[f'A{row+1}'] = "Class:"; ws[f'B{row+1}'] = class_data["Class"].upper()
    ws[f'D{row+1}'] = "Section:"; ws[f'E{row+1}'] = student["Section"].upper()
    ws[f'G{row+1}'] = "No in Class:"; ws[f'I{row+1}'] = len(students_data)
    ws[f'A{row+2}'] = "Resumption:"; ws[f'B{row+2}'] = class_data["Resumption Date"]

    for r in range(row, row+3):
        for c in range(1, 14):
            cell = ws.cell(r, c)
            cell.font = title_font
            cell.border = border
            cell.alignment = right_align if c in [1,4,7] else center_align

    row += 4

    # Academic Table
    ws.merge_cells(f'A{row}:M{row}')
    ws[f'A{row}'] = "ACADEMIC PERFORMANCE"
    ws[f'A{row}'].font = sub_header_font
    ws[f'A{row}'].fill = gray_fill
    ws[f'A{row}'].alignment = center_align
    ws.row_dimensions[row].height = 25
    row += 1

    table_start = row

    # Define columns based on term
    if term_num == 1:
        headers = ["SUBJECT", "CA1 (20%)", "CA2 (20%)", "CA3 (20%)", "EXAM(40%)", "TOTAL", "AVERAGE", "GRADE", "POSITION", "REMARK"]
        cols_total = 6
        show_1st = False
        show_2nd = False
    elif term_num == 2:
        headers = ["SUBJECT", "CA1(20%)", "CA2(20%)", "CA3(20%)", "EXAM(40%)", "TOTAL", "1ST TERM", "CUMULATIVE", "GRADE", "POSITION", "REMARK"]
        cols_total = 6
        show_1st = True
        show_2nd = False
    else:  # 3rd Term
        headers = ["SUBJECT", "CA1(20%)", "CA2(20%)", "CA3(20%)", "EXAM(40%)", "TOTAL", "1ST TERM", "2ND TERM", "CUMULATIVE", "GRADE", "POSITION", "REMARK"]
        cols_total = 6
        show_1st = True
        show_2nd = True

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row, i, h)
        cell.font = title_font
        cell.alignment = center_align
        cell.fill = gray_fill
        cell.border = border
    ws.row_dimensions[row].height = 30
    # row += 1
   # Merge position/remark columns
    if term_num == 1:
        print("row J", row)
        ws.merge_cells(f'J8:M8')
    elif term_num == 2:
        print("row S", row)
        ws.merge_cells(f'k8:M8')
        print("row T", row)
    else:
        ws.merge_cells(f'L8:M8')
    

    # Subject rows
    for idx, subject in enumerate(subjects):
        r = row+1+ idx
        # Set subject name in column A
        cell = ws.cell(r, 1)
        cell.value = subject.upper()
        cell.alignment = Alignment(horizontal="left")
        cell.font = title_font
        cell.border = border
        
        # CA1 to Exam (will be filled later or left as 0)
        for c in range(2, 6):
            ws.cell(r, c, 0)

        # This Term Total
        ws.cell(r, cols_total).value = f"=ROUND(B{r} + C{r} + D{r} + E{r}, 1)"

        col_offset = 7

        # 1st Term Score (only show on 2nd & 3rd term)
        if term_num >= 2 and show_1st:
            sheet_1st = safe_sheet_name(student["Name"].upper().replace(" ", "_"), "1ST_TERM")
            ws.cell(r, col_offset).value = f"='{sheet_1st}'!F{r}" if term_num >= 2 else ""
            col_offset += 1

        # 2nd Term Score (only on 3rd term)
        if term_num == 3 and show_2nd:
            sheet_2nd = safe_sheet_name(student["Name"].upper().replace(" ", "_"), "2ND_TERM")
            ws.cell(r, col_offset).value = f"='{sheet_2nd}'!F{r}"
            col_offset += 1

        # Cumulative Total
        if term_num == 1:
            cum_col = col_offset
            ws.cell(r, cum_col).value = f"=F{r}"
        elif term_num == 2:
            cum_col = col_offset
            ws.cell(r, cum_col).value = f"=F{r} + G{r}"
        else:
            cum_col = col_offset
            ws.cell(r, cum_col).value = f"=F{r} + G{r} + H{r}"

        # Grade (based on this term total)
        grade_col = cum_col + 1
        ws.cell(r, grade_col).value = f'=IF(F{r}>=75,"A",IF(F{r}>=60,"B",IF(F{r}>=50,"C",IF(F{r}>=45,"D",IF(F{r}>=40,"E","F")))))'

        # Remark
        remark_col = grade_col + 2
        ws.cell(r, remark_col).value = f'=IF({ws.cell(r, grade_col).coordinate}="A","Excellent",IF({ws.cell(r, grade_col).coordinate}="B","Very Good",IF({ws.cell(r, grade_col).coordinate}="C","Good",IF({ws.cell(r, grade_col).coordinate}="D","Fair",IF({ws.cell(r, grade_col).coordinate}="E","Weak","Poor")))))'
        
        # Merge remark cell to column M
        ws.merge_cells(f'{ws.cell(r, remark_col).coordinate}:M{r}')
        for c in range(remark_col, 14):
            ws.cell(r, c).border = border

        # Apply styling
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.alignment = center_align
            cell.border = border

    row += len(subjects) + 2

    # Summary Section
    summary_row = table_start + len(subjects) + 2
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "SUMMARY"
    ws[f'A{row}'].font = sub_header_font
    ws[f'A{row}'].fill = gray_fill

    total_obtainable = len(subjects) * 100
    total_row = row + 1
    ws[f'A{total_row}'] = "Total Score Obtainable"
    ws[f'B{total_row}'] = total_obtainable

    ws[f'A{total_row+1}'] = "Total Score Obtained"
    ws[f'B{total_row+1}'] = f"=SUM(F{table_start+1}:F{row-2})"

    ws[f'A{total_row+2}'] = "Average"
    ws[f'B{total_row+2}'] = f"=ROUND(AVERAGE(F{table_start+1}:F{row-2}),1)"

    ws[f'A{total_row+3}'] = "Overall Grade"
    ws[f'B{total_row+3}'] = f'=IF(B{total_row+2}>=75,"A",IF(B{total_row+2}>=60,"B",IF(B{total_row+2}>=50,"C",IF(B{total_row+2}>=45,"D",IF(B{total_row+2}>=40,"E","F")))))'

    ws[f'A{total_row+4}'] = "Position in Class"
    ws[f'B{total_row+4}'] = "________"

    for i in range(5):
        ws[f'A{total_row+i}'].font = title_font
        ws[f'B{total_row+i}'].font = title_font
        ws[f'B{total_row+i}'].alignment = Alignment(horizontal="right")
        ws[f'A{total_row+i}'].border = border
        ws[f'B{total_row+i}'].border = border


    # Chart (based on term)
    chart_row = total_row + 6
    chart = BarChart()
    chart.title = f"{term_name.upper()} TERM PERFORMANCE"
    chart.y_axis.title = "Score"
    chart.x_axis.title = "Subjects"
    chart.height = 9
    chart.width = 12
    chart.style = 12
    chart.legend = None
    chart.x_axis.majorTickMark = "in"
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    # Adjust data column based on term (use cumulative total)
    if term_num == 1:
        data_col = 6  # Column E (This Term Total)
    elif term_num == 2:
        data_col = 6  # Column F (Cumulative for 2nd term)
    else:  # term_num == 3
        data_col = 6  # Column G (Cumulative for 3rd term)

    data = Reference(ws, min_col=data_col, min_row=table_start, max_row=table_start + len(subjects) - 1)
    cats = Reference(ws, min_col=1, min_row=table_start + 1, max_row=table_start + len(subjects))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.x_axis.textRotation = 45
    chart.x_axis.title.txPr = None
    chart.y_axis.title.txPr = None
    chart.title.txPr = None
    chart.x_axis.title.font = Font(size=12, bold=True)
    chart.y_axis.title.font = Font(size=12, bold=True)
    chart.title.font = Font(size=14, bold=True)
    chart.gapWidth = 150

    ws.add_chart(chart, f"A{chart_row}")

    # --- 6. Affective Domain ---
    affective_row = summary_row
    ws[f"D{affective_row}"] = "Effort".upper()
    ws[f"D{affective_row}"].font = title_font
    ws[f"D{affective_row}"].fill = section_fill
    ws[f"D{affective_row}"].alignment = center_align
    ws[f"D{affective_row}"].border = border
    affective_traits = ["Aesthetic Appreciation", "Attendance", "Creativity", "Honesty", "Initiative","Leadership","Neatness","Obedience", "Perserverance","Politeness","Self Control",  "Sense Of Responsibility", "Sociability", "Spirit Of Coordination"]
    for col in range(5):
        ws.cell(row=affective_row, column=5 + col).value = col + 1
        ws.cell(row=affective_row, column=5 + col).alignment = center_align
        ws.cell(row=affective_row, column=5 + col).font = title_font
        ws.cell(row=affective_row, column=5 + col).border = border
    for i, trait in enumerate(affective_traits):
        ws[f"D{affective_row+1+i}"] = trait.upper()
        ws[f"E{affective_row+1+i}"] = "✓"
        ws[f"D{affective_row+1+i}"].border = border
        for col in range(5):
            ws.cell(row=affective_row+1+i, column=5 + col).border = border


    # --- 7. Psychomotor Skills ---
    psy_row = affective_row + len(affective_traits) + 2
    ws[f"D{psy_row}"] = "Psychomotor Skills".upper()
    ws[f"D{psy_row}"].font = title_font
    ws[f"D{psy_row}"].fill = section_fill
    ws[f"D{psy_row}"].alignment = center_align
    ws[f"D{psy_row}"].border = border
    for col in range(5):
        ws.cell(row=psy_row, column=5 + col).value = col + 1
        ws.cell(row=psy_row, column=5 + col).alignment = center_align
        ws.cell(row=psy_row, column=5 + col).font = title_font
        ws.cell(row=psy_row, column=5 + col).border = border
    psy_traits = ["Communication Skills", "Craft", "Games/Sports", "Handwriting", "Handling Of Tools", "Musical Skills", "Painting/Drawing"]
    for i, trait in enumerate(psy_traits):
        ws[f"D{psy_row+1+i}"] = trait
        ws[f"E{psy_row+1+i}"] = "✓"
        ws[f"D{psy_row+1+i}"].border = border
        for col in range(5):
            ws.cell(row=psy_row+1+i, column=5 + col).border = border

    # --- 8. Remarks & Signature Section ---
    remark_row = psy_row + len(psy_traits) + 2
    ws[f"A{remark_row}"] = "Teacher's Remark:".upper()
    ws[f"A{remark_row+1}"] = f"{class_data['Coordinator']}'s Remark:".upper()
    ws[f"A{remark_row+2}"] = "Date:".upper()
    ws[f"B{remark_row+2}"] = datetime.today().strftime('%d-%m-%Y')
    for i in range(3):
        ws[f"A{remark_row+i}"].border = border
        ws[f"A{remark_row+i}"].font = title_font
        ws[f"A{remark_row+i}"].alignment = Alignment(horizontal="left")
        ws[f"A{remark_row+i}"].fill = section_fill
        ws.row_dimensions[remark_row + i].height = 22
        ws[f"B{remark_row+i}"].border = border
        ws[f"B{remark_row+i}"].alignment = Alignment(horizontal="left")
        ws[f"B{remark_row+i}"].font = title_font
        ws[f"B{remark_row+i}"].fill = section_fill
        ws[f"B{remark_row+i}"].value = "To be filled".upper()
        ws[f"C{remark_row+i}"].border = border

        ws.merge_cells(start_row=remark_row + i, start_column=1, end_row=remark_row + i, end_column=13)

    # --- 9. Grading Scale ---
    grade_row = summary_row
    ws[f"K{grade_row}"] = "Grading Scale (Legend)".upper()
    ws[f"K{grade_row}"].font = title_font
    ws[f"K{grade_row}"].fill = section_fill
    ws[f"K{grade_row}"].alignment = center_align
    ws[f"K{grade_row}"].border = border
    ws.row_dimensions[grade_row].height = 25
    # ws.merge_cells("K{grade_row}:L{grade_row}")
    ws.merge_cells(f"K{grade_row}:L{grade_row}")
    grades = {
        "A": "75 - 100 (Excellent)",
        "B": "60 - 74 (Very Good)",
        "C": "50 - 59 (Good)",
        "D": "45 - 49 (Fair)",
        "E": "40 - 44 (Weak)",
        "F": "0 - 39 (Poor)"
    }
    for i, (grade, meaning) in enumerate(grades.items()):
        ws[f"K{grade_row+1+i}"] = grade
        ws[f"L{grade_row+1+i}"] = meaning.upper()
        ws[f"K{grade_row+1+i}"].border = border
        ws[f"L{grade_row+1+i}"].border = border

    # --- Page Setup for Printing ---
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1  # Fit to one page wide
    ws.page_setup.fitToHeight = 0  # Let data stretch vertically (no limit)
    ws.page_setup.horizontalCentered = True
    ws.page_setup.verticalCentered = False
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2

# === CREATE SHEETS FOR EACH STUDENT ===
for student in students_data:
    base = student["Name"].upper().replace(" ", "_")

    # 1st Term
    ws1 = wb.create_sheet(safe_sheet_name(base, "1ST_TERM"))
    create_term_sheet(ws1, student, "First", 1, show_prev=False)

    # 2nd Term
    ws2 = wb.create_sheet(safe_sheet_name(base, "2ND_TERM"))
    create_term_sheet(ws2, student, "Second", 2, show_prev=True)

    # 3rd Term (full cumulative)
    ws3 = wb.create_sheet(safe_sheet_name(base, "3RD_TERM"))
    create_term_sheet(ws3, student, "Third", 3, show_prev=True)

# Save
os.makedirs("final_3term_results", exist_ok=True)
filename = f"{class_data['Class']}_COMPLETE_3-TERM_RESULTS_2024_2025.xlsx"
path = os.path.join("final_3term_results", filename)
wb.save(path)

print("SUCCESS! 3-Term Report Card Generated")
print(f"File: {path}")
print(f"Students: {len(students_data)} × 3 sheets = {len(students_data)*3} total sheets")
print("1st Term → Only current term")
print("2nd Term → Current + 1st Term")
print("3rd Term → All 3 terms + Cumulative")