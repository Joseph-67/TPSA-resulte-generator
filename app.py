import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
# --- Sample student data for one student offering 14 subjects ---
from classes.grade.subject import subjects
from classes.grade.gradeOneDiamond import students_data, class_data
student_scores = [
    {"Subject": sub, "CA1": 0, "CA2": 0, "CA3": 0, "Exam": 0,
     "3rd Term": 0, "1st Term": 0, "2nd Term": 0} for sub in subjects
]

# Convert student scores to DataFrame for easier manipulation
df = pd.DataFrame(student_scores)

wb = Workbook()
wb.remove(wb.active)
for student in students_data:
    name = student["Name"]
    style_sheet_name = name if len(name) <= 31 else name[:28]
    ws = wb.create_sheet(title=style_sheet_name)

    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.title = student["Name"].upper()

    # --- Set column widths for neat layout ---
    col_widths = [18, 7, 7, 7, 7, 13, 10, 10, 13, 8, 8, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    # --- Styling ---
    header_font = Font(name="Times New Roman", bold=True, size=36, shadow=True)
    sub_header_font = Font(name="Times New Roman", bold=True, size=16, shadow=True)
    title_font = Font(name="Times New Roman", bold=True, size=12, shadow=True)
    section_fill = PatternFill("solid", fgColor="DDDDDD")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- 1. Sheet Title & Header ---
    ws.merge_cells('A1:M1')
    ws['A1'] = "TOPSTEPS ACADEMY"
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30

    # --- Sub Header Section ---
    ws.merge_cells('A2:M2')
    ws['A2'] = "SECOND TERM 2024/2025 SESSION ACADEMIC REPORT"
    ws['A2'].font = sub_header_font
    ws['A2'].alignment = center_align
    ws['A2'].fill = section_fill
    ws.row_dimensions[2].height = 22

    # Insert logo
    logo = Image("tpsa_logo.png")
    logo.width, logo.height = 70, 70
    ws.add_image(logo, "M1")

    # --- 2. Student Bio Data ---


    ws.merge_cells('B3:F3')
    ws.merge_cells('G3:H3')
    ws.merge_cells('I3:M3')
    ws.merge_cells('B4:C4')
    ws.merge_cells('E4:F4')
    ws.merge_cells('G4:H4')
    ws.merge_cells('I4:M4')
    ws.merge_cells('B5:M5')
    ws.merge_cells('A7:M7')

    ws['A3'] = "Student Name: ".upper()
    ws['B3'] = student["Name"].upper()
    ws['A3'].font = title_font
    ws['B3'].font = title_font
    ws['A3'].alignment = Alignment(horizontal="right")
    ws['B3'].alignment = center_align
    ws['A3'].border = border
    ws['B3'].border = border
    ws['C3'].border = border
    ws['D3'].border = border
    ws['E3'].border = border
    ws['F3'].border = border
    ws['G3'] = "Admission No: ".upper()
    ws['I3'] = student["Admission No"]
    ws['G3'].font = title_font
    ws['I3'].font = title_font
    ws['G3'].alignment = Alignment(horizontal="right")
    ws['I3'].alignment = center_align
    ws['G3'].border = border
    ws['I3'].border = border
    ws['H3'].border = border
    ws['J3'].border = border
    ws['K3'].border = border
    ws['L3'].border = border
    ws['M3'].border = border
    ws['A4'] = "Class: ".upper()
    ws['B4'] = class_data["Class"].upper()
    ws['A4'].font = title_font
    ws['B4'].font = title_font
    ws['A4'].alignment = Alignment(horizontal="right")
    ws['B4'].alignment = center_align
    ws['A4'].border = border
    ws['B4'].border = border
    ws['C4'].border = border
    ws['D4'] = "Section: ".upper()
    ws['E4'] = student["Section"].upper()
    ws['D4'].font = title_font
    ws['E4'].font = title_font
    ws['D4'].alignment = Alignment(horizontal="right")
    ws['E4'].alignment = center_align
    ws['D4'].border = border
    ws['E4'].border = border
    ws['G4'] = "Number In Class: ".upper()
    ws['I4'] = len(students_data)
    ws['G4'].font = title_font
    ws['I4'].font = title_font
    ws['G4'].alignment = Alignment(horizontal="right")
    ws['I4'].alignment = center_align
    ws['G4'].border = border
    ws['I4'].border = border
    ws['H4'].border = border
    ws['J4'].border = border
    ws['K4'].border = border
    ws['L4'].border = border
    ws['M4'].border = border
    ws['A5'] = "Resumption Date: ".upper()
    ws['B5'] = class_data["Resumption Date"]
    ws['A5'].font = title_font
    ws['B5'].font = title_font
    ws['A5'].alignment = Alignment(horizontal="right")
    ws['B5'].alignment = center_align
    ws['A5'].border = border
    ws['B5'].border = border
    ws['C5'].border = border
    ws['D5'].border = border
    ws['E5'].border = border
    ws['F5'].border = border
    ws['G5'].border = border
    ws['H5'].border = border
    ws['I5'].border = border
    ws['J5'].border = border
    ws['K5'].border = border
    ws['L5'].border = border
    ws['M5'].border = border
    # Prepare keys in pairs
    items = list(student.items())
    row = 3
    col_widths = {"A": 40.71, "B": 15, "C": 15, "D": 25, "E" : 15, "F": 15, "G": 15, "H": 15, "I": 15, "J": 16, "K" : 15, "L": 20, "M": 30}

    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    for i in range(0, len(items), 2):
        ws.row_dimensions[row].height = 20
        row += 1
    # --- 3. Academic Performance Table ---
    ws['A7'] = "Academic Performance".upper()
    ws['A7'].font = sub_header_font
    ws['A7'].fill = section_fill
    ws['A7'].alignment = center_align
    ws.row_dimensions[7].height = 22
    table_start_row = row + 1
    columns = [
        "Subject", "CA1 (20%)", "CA2 (20%)", "CA3 (20%)", "Exam (40%)",
        "3rd Term Total", "Class Average", "1st Term", "2nd Term", "Cumulative Total",
        "Grade", "Position", "Remark"
    ]
    for col_num, col_name in enumerate(columns, 1):
        cell = ws.cell(row=table_start_row, column=col_num, value=col_name.upper())
        cell.font = title_font
        cell.alignment = center_align
        cell.fill = section_fill
        cell.border = border
    ws.row_dimensions[table_start_row].height = 30

    for i, subject in enumerate(subjects):
        r = table_start_row + 1 + i
        row_data = df.iloc[i]
        ws.cell(row=r, column=1).value = row_data["Subject"].upper()
        ws.cell(row=r, column=2).value = row_data["CA1"]
        ws.cell(row=r, column=3).value = row_data["CA2"]
        ws.cell(row=r, column=4).value = row_data["CA3"]
        ws.cell(row=r, column=5).value = row_data["Exam"]
        # 3rd Term Total: weighted sum
        ws.cell(row=r, column=6).value = f"=ROUND(B{r}*0.2+C{r}*0.2+D{r}*0.2+E{r}*0.4, 1)"
        # Class Average (dummy, can be replaced with actual average if available)
        ws.cell(row=r, column=7).value = None  # Placeholder for class average
        # 1st Term and 2nd Term scores (dummy, can be replaced with
        ws.cell(row=r, column=8).value = row_data["1st Term"]
        ws.cell(row=r, column=9).value = row_data["2nd Term"]
        # Cumulative Total: sum of all terms
        ws.cell(row=r, column=10).value = f"=F{r}+H{r}+I{r}"
        # Grade based on 3rd Term Total
        ws.cell(row=r, column=11).value = f'''=IF(F{r}>=75,"A",IF(F{r}>=60,"B",IF(F{r}>=50,"C",IF(F{r}>=45,"D",IF(F{r}>=40,"E","F")))))'''
        ws.cell(row=r, column=12).value = ""  # Position (to be filled later)
        ws.cell(row=r, column=13).value = f'''=IF(K{r}="A","Excellent",IF(K{r}="B","Very Good",IF(K{r}="C","Good",IF(K{r}="D","Fair",IF(K{r}="E","Weak","Poor")))))'''
        for c in range(1, 14):
            ws.cell(row=r, column=c).alignment = center_align
            ws.cell(row=r, column=c).border = border

    # --- 4. Summary Section ---
    summary_row = table_start_row + len(subjects) + 2
    ws[f"A{summary_row}"] = "Summary".upper()
    ws[f"A{summary_row}"].font = title_font
    ws[f"A{summary_row}"].fill = section_fill
    ws[f"A{summary_row}"].alignment = center_align
    ws[f"A{summary_row}"].border = border
    ws[f"B{summary_row}"].border = border
    ws[f"A{summary_row}"].font = sub_header_font
    ws.row_dimensions[summary_row].height = 20
    ws.merge_cells(f"A{summary_row}:B{summary_row}")

    ws[f"A{summary_row+1}"] = "Total Score".upper()
    ws[f"B{summary_row+1}"] = f'=ROUND(SUM(F{table_start_row+1}:F{table_start_row+len(subjects)}), 1)'
    ws[f"A{summary_row+2}"] = "Average Score".upper()
    ws[f"B{summary_row+2}"] = f"=ROUND(AVERAGE(F{table_start_row+1}:F{table_start_row+len(subjects)}), 1)"
    ws[f"A{summary_row+3}"] = "Grade".upper()
    ws[f"B{summary_row+3}"] = f'''=IF(B{summary_row+2}>=75,"A",IF(B{summary_row+2}>=60,"B",IF(B{summary_row+2}>=50,"C",IF(B{summary_row+2}>=45,"D",IF(B{summary_row+2}>=40,"E","F")))))'''
    ws[f"A{summary_row+4}"] = "Position in Class".upper()
    ws[f"B{summary_row+4}"] = "To be filled"
    for i in range(1, 5):
        ws[f"A{summary_row+i}"].border = border
        ws[f"B{summary_row+i}"].border = border
        ws[f"A{summary_row+i}"].font = title_font
        ws[f"B{summary_row+i}"].font = title_font
        ws[f"B{summary_row+i}"].alignment = Alignment('right')

    # --- 5. Graph (Performance Chart) ---
    # Place chart below summary section
    chart_row = summary_row + 7

    # Prepare data for chart (exclude header for categories)
    data = Reference(ws, min_col=6, min_row=table_start_row, max_row=table_start_row + len(subjects))
    # Explicitly set all subjects as categories for x-axis
    cats = Reference(ws, min_col=1, min_row=table_start_row+1, max_row=table_start_row + len(subjects))

    chart = BarChart()
    # chart.type = "col"
    chart.title = "Academic Performance (Third Term)"
    chart.x_axis.title = "Subjects"
    chart.y_axis.title = "Total Score"
    chart.height = 9
    chart.width = 12
    chart.style = 1
    # chart.style = 10
    chart.legend = None  # Hide legend for clarity
    chart.y_axis.majorGridlines = None  # Hide gridlines for cleaner look
    chart.x_axis.majorGridlines = None  # Hide x-axis gridlines
    chart.x_axis.majorTickMark = "in"  # Set x-axis ticks to be inside
    # Add data to chart


    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Ensure all subject names are shown on x-axis
    chart.x_axis.majorTickMark = "out"
    chart.x_axis.tickLblSkip = 1  # Show every label

    # chart.dataLabels = DataLabelList()
    # chart.dataLabels.showVal = True
    # Rotate subject labels for readability
    chart.x_axis.textRotation = 45

    # Improve chart layout: center chart, increase spacing, adjust axis and title font sizes
    chart.layout = Layout(
        manualLayout=ManualLayout(
            x=0.0,  # Center horizontally (0.0 = left, 1.0 = right)
            y=0.05,  # Move a bit down from the top
            h=0.9,   # Height as fraction of plot area
            w=0.9    # Width as fraction of plot area
        )
    )
    chart.x_axis.title.txPr = None  # Remove text properties to use default
    chart.y_axis.title.txPr = None
    chart.title.txPr = None
    chart.x_axis.title.font = Font(size=12, bold=True)
    chart.y_axis.title.font = Font(size=12, bold=True)
    chart.title.font = Font(size=14, bold=True)
    chart.gapWidth = 150  # Increase gap between bars for clarity

    # Explicitly set x and y axis labels (redundant, but ensures labels are set)
    chart.x_axis.title = "Subjects"
    chart.y_axis.title = "Total Score"

    # Insert chart
    ws.add_chart(chart, f"A{chart_row}")

    # --- 6. Affective Domain ---
    affective_row = summary_row
    ws[f"D{affective_row}"] = "Effort".upper()
    ws[f"D{affective_row}"].font = title_font
    ws[f"D{affective_row}"].fill = section_fill
    ws[f"D{affective_row}"].alignment = center_align
    ws[f"D{affective_row}"].border = border
    affective_traits = ["Aesthetic Appreciation", "Attendance", "Creativity", "Honesty", "Initiative","Leadership","Neatness","Obedience", "Perserverance","Politeness","Self Control",  "Sense Of Responsibility", "Sociability", "Spirit Of Coordination"]
    for col in range(5):
        ws.cell(row=affective_row, column=5 + col).value = col + 1
        ws.cell(row=affective_row, column=5 + col).alignment = center_align
        ws.cell(row=affective_row, column=5 + col).font = title_font
        ws.cell(row=affective_row, column=5 + col).border = border
    for i, trait in enumerate(affective_traits):
        ws[f"D{affective_row+1+i}"] = trait.upper()
        ws[f"E{affective_row+1+i}"] = "✓"
        ws[f"D{affective_row+1+i}"].border = border
        for col in range(5):
            ws.cell(row=affective_row+1+i, column=5 + col).border = border


    # --- 7. Psychomotor Skills ---
    psy_row = affective_row + len(affective_traits) + 2
    ws[f"D{psy_row}"] = "Psychomotor Skills".upper()
    ws[f"D{psy_row}"].font = title_font
    ws[f"D{psy_row}"].fill = section_fill
    ws[f"D{psy_row}"].alignment = center_align
    ws[f"D{psy_row}"].border = border
    for col in range(5):
        ws.cell(row=psy_row, column=5 + col).value = col + 1
        ws.cell(row=psy_row, column=5 + col).alignment = center_align
        ws.cell(row=psy_row, column=5 + col).font = title_font
        ws.cell(row=psy_row, column=5 + col).border = border
    psy_traits = ["Communication Skills", "Craft", "Games/Sports", "Handwriting", "Handling Of Tools", "Musical Skills", "Painting/Drawing"]
    for i, trait in enumerate(psy_traits):
        ws[f"D{psy_row+1+i}"] = trait
        ws[f"E{psy_row+1+i}"] = "✓"
        ws[f"D{psy_row+1+i}"].border = border
        for col in range(5):
            ws.cell(row=psy_row+1+i, column=5 + col).border = border

    # --- 8. Remarks & Signature Section ---
    remark_row = psy_row + len(psy_traits) + 2
    ws[f"A{remark_row}"] = f"{class_data['Coordinator']}'s Remark:".upper()
    ws[f"A{remark_row+1}"] = "Parent/Guardian's Remark:".upper()
    ws[f"A{remark_row+2}"] = "Date:".upper()
    ws[f"B{remark_row+2}"] = datetime.today().strftime('%d-%m-%Y')
    for i in range(3):
        ws[f"A{remark_row+i}"].border = border
        ws[f"A{remark_row+i}"].font = title_font
        ws[f"A{remark_row+i}"].alignment = Alignment(horizontal="left")
        ws[f"A{remark_row+i}"].fill = section_fill
        ws.row_dimensions[remark_row + i].height = 22
        ws[f"B{remark_row+i}"].border = border
        ws[f"B{remark_row+i}"].alignment = Alignment(horizontal="left")
        ws[f"B{remark_row+i}"].font = title_font
        ws[f"B{remark_row+i}"].fill = section_fill
        ws[f"B{remark_row+i}"].value = "To be filled".upper()
        ws[f"C{remark_row+i}"].border = border

        ws.merge_cells(start_row=remark_row + i, start_column=1, end_row=remark_row + i, end_column=13)

    # --- 9. Grading Scale ---
    grade_row = summary_row
    ws[f"K{grade_row}"] = "Grading Scale (Legend)".upper()
    ws[f"K{grade_row}"].font = title_font
    ws[f"K{grade_row}"].fill = section_fill
    ws[f"K{grade_row}"].alignment = center_align
    ws[f"K{grade_row}"].border = border
    ws.row_dimensions[grade_row].height = 25
    # ws.merge_cells("K{grade_row}:L{grade_row}")
    ws.merge_cells(f"K{grade_row}:L{grade_row}")
    grades = {
        "A": "75 - 100 (Excellent)",
        "B": "60 - 74 (Very Good)",
        "C": "50 - 59 (Good)",
        "D": "45 - 49 (Fair)",
        "E": "40 - 44 (Weak)",
        "F": "0 - 39 (Poor)"
    }
    for i, (grade, meaning) in enumerate(grades.items()):
        ws[f"K{grade_row+1+i}"] = grade
        ws[f"L{grade_row+1+i}"] = meaning.upper()
        ws[f"K{grade_row+1+i}"].border = border
        ws[f"L{grade_row+1+i}"].border = border

    # --- Page Setup for Printing ---
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1  # Fit to one page wide
    ws.page_setup.fitToHeight = 0  # Let data stretch vertically (no limit)
    ws.page_setup.horizontalCentered = True
    ws.page_setup.verticalCentered = False
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

folder_path = "template"
os.makedirs(folder_path, exist_ok=True)

wb.save(os.path.join(folder_path, f"{class_data['Class']}_Report_Card_Template.xlsx"))
