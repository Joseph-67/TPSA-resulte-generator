import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.cell.text import InlineFont
from openpyxl.cell.text import RichText
# from PIL import Image
from openpyxl.drawing.image import Image

from openpyxl.chart import BarChart, Series, Reference, BarChart3D
import studentlist as ST
import os
from io import BytesIO
# create a workbook
wb = Workbook()

# load main template
wb = load_workbook("templates/2024_2025/second_term/preschool TEMPLATE V1.0 (2).xlsx")
# create a worksheet
ws = wb.active
print(ws["B3"].value)

list_of_students = ST.preNurseryI

# names = [ ]
no_of_students = len(list_of_students)
count = 1
for x in list_of_students:
    sheet_to_copy = wb['Sheet1']
    new_sheet = wb.copy_worksheet(sheet_to_copy)
    new_sheet.title = x[:12]
    # number of students

    new_sheet["B3"].value = x.upper()
    print(new_sheet["B3"].value)
    new_sheet["B5"].value = "PRESCHOOL ONE"
    new_sheet["F5"].value = ""
    new_sheet["J5"].value = no_of_students
    
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "SECOND TERM ACADEMIC PERFORMANCE"
    chart1.y_axis.title = 'Total Score'
    chart1.x_axis.title = 'List of Subjects'
    chart1.height = 20
    # chart1.height = 15
    chart1.width = 14.5

    data = Reference(new_sheet, min_col=7, min_row=9, max_row=18)
    cats = Reference(new_sheet, min_col=1, min_row=10, max_row=18)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    # chart1.shape = 4
    new_sheet.add_chart(chart1, "A29")
    app_name = 'PRESCHOOL1'

    openpyxl_version = openpyxl.__version__
    # print(openpyxl_version)

    # logo insertion begins here
    img_file = 'logo.png'
    # with open("logo.png", "rb") as img_file:
        # img_file.write(b"ResultGenerator/logo.png")
    img = openpyxl.drawing.image.Image(img_file)
    # img.anchor = 'A1'
    img.width = 150.21
    img.height = 150.69
    new_sheet.add_image(img, "A1")
        # ends here

    # heading
    new_sheet["B1"].value = "TOPSTEPS ACADEMY"
    new_sheet["B1"].font = Font(name="Times New Roman", sz= 72, b=True, shadow=True)
    new_sheet["B2"].value = "SECOND TERM 2024/2025 SESSION ACADEMIC REPORT"
    new_sheet["B2"].font = Font(name="Times New Roman", sz= 28, b=True, shadow=True)
    # end heading
    

    wb.save(f"Reports/2024-2025/second-term/{app_name}.xlsx")